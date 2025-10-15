import os
import random
import pandas as pd
from faker import Faker
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ─── CONFIG ─────────────────────────────────────────────────────────────────────
excel_dir   = r"C:\Users\kjden\source\repos\Kieshdh\KieshStockExchange"
excel_name  = "AIUserData.xlsx"
excel_path  = os.path.join(excel_dir, excel_name)
sheet_name  = "Data"
NUM_PEOPLE  = 1000
fake = Faker()
stock_tickers = ["MSFT", "NVDA", "AAPL", "AMZN", "GOOG", "META", "AVGO", "TSLA", "TSM", "WMT",
                "LLY", "V", "ORCL", "NFLX", "MA", "BAC", "ASML", "KO", "BABA", "MCD", "AMD"]
stock_prices = [513.71, 173.50, 213.88, 231.44, 194.08, 712.68, 290.18, 316.06, 245.6, 97.47, 812.69, 
                357.04, 245.12, 1180.49, 568.22, 48.45, 711.25, 69.17, 120.03, 298.47, 166.47]
# ────────────────────────────────────────────────────────────────────────────────

wb = load_workbook(excel_path)
if 'Blad1' in wb.sheetnames:
    std = wb['Blad1']
    wb.remove(std)
ws_identity = wb.create_sheet("Identity")
#ws_profile  = wb.create_sheet("Profile")
ws_preference = wb.create_sheet("Preference")
ws_holding = wb.create_sheet("Holding")
#ws_stocks   = wb.create_sheet("Stocks")

# Sheet headers
#ws_identity.append(["ID", "Username", "Full Name", "Email", "Birthdate"])
#ws_profile.append(["ID", "Min Stocks", "Max Stocks", "Watchlist Size", "Online Score", "Aggressive", "Balance"])
ws_preference.append(["ID"] + stock_tickers)
ws_holding.append(["ID", "Balance"] + stock_tickers)

# Define dark theme styles
header_fill = PatternFill(start_color="2E2E2E", end_color="2E2E2E", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
row_fill_alt = PatternFill(start_color="1E1E1E", end_color="1E1E1E", fill_type="solid")
text_font = Font(color="FFFFFF")

def apply_dark_theme(ws):
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Style rest of the rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = text_font
            cell.fill = row_fill_alt

def beta_biased_int(low, high, a=0.8, b=3.0):
    span = high - low + 1
    r = random.betavariate(a, b)
    return low + int(r * span)

fake = Faker()

def generate_username():
    while True:
        username = fake.user_name()
        if len(username) >= 5:
            return username

class Person:
    def __init__(self):
        # identity
        self.full_name      = fake.name()
        self.username       = generate_username()
        self.email          = f"{self.username}@{fake.free_email_domain()}"
        self.birthdate      = fake.date_of_birth(minimum_age=18, maximum_age=80)

        # skewed‐toward‐low
        #  ─ min_stocks: 1–8
        self.min_stocks = beta_biased_int(2, 8)
        #  ─ max_stocks: at least min_stocks+1, up to 15
        min_max = max(self.min_stocks + 1, 3)
        self.max_stocks = beta_biased_int(min_max, 15)
        #  ─ watchlist_size: > max_stocks, up to 20
        self.watchlist_size = beta_biased_int(self.max_stocks + 1, 20)
        #  ─ online_score: 5–100%
        self.online_score = beta_biased_int(5, 100, 0.5, 3)/100
        #  ─ aggressive: 10–100%
        self.aggressive = beta_biased_int(10, 100, 2, 2.5)/100
        self.balance = beta_biased_int(1000, 1000000)

    def __repr__(self):
        return (
            f"Person({self.username!r}, min={self.min_stocks}, max={self.max_stocks}, "
            f"watch={self.watchlist_size}, online={self.online_score}%, "
            f"aggr={self.aggressive}%)"
        )

# ─── GENERATE & WRITE ────────────────────────────────────────────────────────────
for idx, p in enumerate((Person() for _ in range(NUM_PEOPLE)), start=1):
    # Identity data
    #ws_identity.append([idx, p.username, p.full_name, p.email, p.birthdate])

    # Profile data
    #ws_profile.append([ idx, p.min_stocks, p.max_stocks, p.watchlist_size, p.online_score, p.aggressive, p.balance ])

    # owned_count between min and max (also skewed low)
    owned_count = random.randint(p.min_stocks, p.max_stocks)
    owned       = set(random.sample(stock_tickers, owned_count))
    extra_needed = p.watchlist_size - owned_count
    watch_only  = set(random.sample(
        [s for s in stock_tickers if s not in owned],
        extra_needed
    ))

    # fill stock cols with 0/1/2
    stock_row = [idx]
    for ticker in stock_tickers:
        if ticker in owned:
            stock_row.append(2)
        elif ticker in watch_only:
            stock_row.append(1)
        else:
            stock_row.append(0)
    ws_preference.append(stock_row)

    # Holding data
    holding_row = [idx, p.balance]
    avgShareValue = p.balance / owned_count if owned_count > 0 else 0
    for ticker in stock_tickers:
        if ticker in owned: 
            holding_row.append( int(avgShareValue / stock_prices[stock_tickers.index(ticker)]) )
        else:
            holding_row.append(0)
    ws_holding.append(holding_row)

# Apply to all sheets
#apply_dark_theme(ws_identity)
#apply_dark_theme(ws_profile)
#apply_dark_theme(ws_stocks)

wb.save(excel_path)
print(f"✅ Generated {NUM_PEOPLE} people with beta-biased attributes.") 
