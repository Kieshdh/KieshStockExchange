# excel_layout.py

import os
from typing import Dict, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Font, PatternFill, GradientFill, Border, Side,
    Alignment, Protection, NamedStyle, Color
)
from openpyxl.utils import get_column_letter

# ────────────────────────────── SHEET STYLING ────────────────────────────────
HEADER_COLOR = Color(rgb="4A7A41")            # DarkGreen
HEADER_FONT_COLOR = Color(rgb="FFFFFF")       # White
ROW_FILL_COLOR = Color(rgb="0F1E25")          # DarkNavy
ROW_ALT_FILL_COLOR = Color(rgb="1E2E35")      # Lighter DarkNavy
TEXT_FONT_COLOR = Color(rgb="FFFFFF")         # White
BORDER_COLOR = Color(rgb="000000")            # Black
BACKGROUND_COLOR = Color(rgb="000000")        # Black

# ────────────────────────────── SHEET CREATION ────────────────────────────────

def load_or_create_workbook(excel_path: str) -> Workbook:
    """Return a fresh workbook for writing at `excel_path`. The file at
    that path is overwritten on save, so any existing contents on disk
    are discarded. The path argument is kept for API compatibility."""
    wb = Workbook()
    wb.active.title = "Template"
    return wb


def reset_or_create_sheet(wb: Workbook, name: str) -> Worksheet:
    """Return a worksheet with the given name that is empty.
    If it already exists, its contents are cleared."""
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)
    return ws


def prepare_stocks_sheet(wb: Workbook) -> Worksheet:
    """Create/reset the Stocks sheet and write its header row."""
    ws = reset_or_create_sheet(wb, "Stocks")
    # Price lives on the Listings sheet (per-currency SeedPrice); not duplicated here.
    ws.append(["StockId", "Ticker", "CompanyName"])
    return ws


def prepare_identity_sheet(wb: Workbook) -> Worksheet:
    """Create/reset the Identity sheet and write its header row."""
    ws = reset_or_create_sheet(wb, "Identity")
    ws.append(["UserId", "Username", "FullName", "Email", "Birthdate", "IsAdmin"])
    return ws


def prepare_holding_sheet(wb: Workbook, tickers) -> Worksheet:
    """Create/reset the Holding sheet and write its header row.
    Includes dynamic ticker columns.
    """
    ws = reset_or_create_sheet(wb, "Holding")
    ws.append(["UserId", "Balance"] + list(tickers))
    return ws


def prepare_profile_sheet(wb: Workbook) -> Worksheet:
    """Create/reset the Profile sheet and write its header row."""
    ws = reset_or_create_sheet(wb, "Profile")
    ws.append([
        "UserId", "Seed", "DecisionIntervalSeconds", "TradeProb",
        "UseMarketProb", "UseSlippageMarketProb",
        "BuyBiasPrc", "MinTradeAmountPrc", "MaxTradeAmountPrc",
        "PerPositionMaxPrc", "MinCashReservePrc", "MaxCashReservePrc",
        "SlippageTolerancePrc", "MinLimitOffsetPrc", "MaxLimitOffsetPrc",
        "AggressivenessPrc",
        "MaxOpenOrders", "WatchlistCsv", "Strategy",
        "ExtremeReactionRandomnessPrc",
        "CashInjectionFrequencyPrc", "CashInjectionAmountPrc",
        "HomeCurrency",
    ])
    return ws


def prepare_listings_sheet(wb: Workbook) -> Worksheet:
    """Create/reset the Listings sheet. One row per (StockId, Currency)."""
    ws = reset_or_create_sheet(wb, "Listings")
    ws.append(["StockId", "Currency", "IsPrimary", "SeedPrice"])
    return ws


def prepare_aiuser_sheets(wb: Workbook, tickers) -> Dict[str, Worksheet]:
    """Create/reset all needed AIUser sheets via dedicated helpers and return them."""
    ws_stocks = prepare_stocks_sheet(wb)
    ws_listings = prepare_listings_sheet(wb)
    ws_identity = prepare_identity_sheet(wb)
    ws_holding = prepare_holding_sheet(wb, tickers)
    ws_profile = prepare_profile_sheet(wb)

    return {
        "Stocks": ws_stocks,
        "Listings": ws_listings,
        "Identity": ws_identity,
        "Holding": ws_holding,
        "Profile": ws_profile,
    }


# ───────────────────────────── THEME / STYLING ────────────────────────────────
def _ensure_dark_theme_styles(wb: Workbook) -> None:
    """Register the dark-theme NamedStyles on the workbook if not present."""
    if "kse_header_edge" in wb.named_styles:
        return

    # Cell styles
    header_fill      = PatternFill(start_color=HEADER_COLOR,       end_color=HEADER_COLOR,       fill_type="solid")
    row_fill_primary = PatternFill(start_color=ROW_FILL_COLOR,     end_color=ROW_FILL_COLOR,     fill_type="solid")
    row_fill_alt     = PatternFill(start_color=ROW_ALT_FILL_COLOR, end_color=ROW_ALT_FILL_COLOR, fill_type="solid")

    # Text styles
    text_font   = Font(color=TEXT_FONT_COLOR)
    header_font = Font(color=HEADER_FONT_COLOR, bold=True)

    # Border styles
    thin_side  = Side(border_style="thin",  color=BORDER_COLOR)
    thick_side = Side(border_style="thick", color=BORDER_COLOR)
    header_edge_border = Border(top=thick_side, bottom=thick_side, left=thin_side, right=thick_side)
    header_mid_border  = Border(top=thick_side, bottom=thick_side, left=thin_side, right=thin_side)
    data_edge_border   = Border(top=thin_side,  bottom=thin_side,  left=thin_side, right=thick_side)
    data_mid_border    = Border(top=thin_side,  bottom=thin_side,  left=thin_side, right=thin_side)

    # Black fill for the dark margin column past the data.
    margin_fill = PatternFill(start_color=BACKGROUND_COLOR, end_color=BACKGROUND_COLOR, fill_type="solid")

    # NamedStyles — six distinct combinations used across the sheet, plus
    # the dark-margin style applied at column level past the data.
    wb.add_named_style(NamedStyle(name="kse_header_edge",    fill=header_fill,      font=header_font, border=header_edge_border))
    wb.add_named_style(NamedStyle(name="kse_header_mid",     fill=header_fill,      font=header_font, border=header_mid_border))
    wb.add_named_style(NamedStyle(name="kse_data_even_edge", fill=row_fill_primary, font=text_font,   border=data_edge_border))
    wb.add_named_style(NamedStyle(name="kse_data_even_mid",  fill=row_fill_primary, font=text_font,   border=data_mid_border))
    wb.add_named_style(NamedStyle(name="kse_data_odd_edge",  fill=row_fill_alt,     font=text_font,   border=data_edge_border))
    wb.add_named_style(NamedStyle(name="kse_data_odd_mid",   fill=row_fill_alt,     font=text_font,   border=data_mid_border))
    wb.add_named_style(NamedStyle(name="kse_dark_margin",    fill=margin_fill))


def apply_dark_theme(ws: Worksheet) -> None:
    """Apply a predefined dark theme to the given worksheet."""
    # Determine worksheet size
    max_used_row = ws.max_row
    max_used_col = ws.max_column
    if max_used_row < 1 or max_used_col < 1:
        return

    _ensure_dark_theme_styles(ws.parent)

    # Per-column style name for each parity — inner loop is then a simple
    # indexed lookup + one attribute set per cell.
    col_even = [
        "kse_data_even_edge" if (c == 1 or c == max_used_col) else "kse_data_even_mid"
        for c in range(1, max_used_col + 1)
    ]
    col_odd = [
        "kse_data_odd_edge" if (c == 1 or c == max_used_col) else "kse_data_odd_mid"
        for c in range(1, max_used_col + 1)
    ]

    # Header row (row 1)
    for c, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_used_col)), start=1):
        cell.style = "kse_header_edge" if (c == 1 or c == max_used_col) else "kse_header_mid"

    # Data rows (2..max_row)
    for r_idx, row in enumerate(
            ws.iter_rows(min_row=2, max_row=max_used_row, min_col=1, max_col=max_used_col),
            start=2):
        # Alternate row styles
        styles = col_even if (r_idx % 2 == 0) else col_odd
        for c_idx, cell in enumerate(row):
            cell.style = styles[c_idx]

    # Dark margin column to the right and dark margin row below the data —
    # keeps the dark theme bleeding past the table edges. Cell-level style
    # is required: Column/RowDimension.style is read-only in this openpyxl
    # version, so each margin cell needs to carry the kse_dark_margin style.
    margin_col_idx = max_used_col + 1
    margin_letter = get_column_letter(margin_col_idx)
    margin_row_idx = max_used_row + 1

    for r in range(1, max_used_row + 1):
        ws.cell(row=r, column=margin_col_idx).style = "kse_dark_margin"
    for c in range(1, margin_col_idx + 1):
        ws.cell(row=margin_row_idx, column=c).style = "kse_dark_margin"

    ws.column_dimensions[margin_letter].width = 150
    ws.row_dimensions[margin_row_idx].height = 200


def autofit_columns(ws: Worksheet, min_width: float = 8.0, max_width: float = 40.0) -> None:
    """
    Approximate auto-fit: set each column width based on the longest cell value
    (in characters) in that column, clamped between min_width and max_width.
    """
    for col_idx, col_values in enumerate(
            ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column,
                         values_only=True),
            start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        for value in col_values:
            if value is None:
                continue

            text = value if isinstance(value, str) else str(value)
            length = len(text)
            if length > max_length:
                max_length = length

        # Skip empty columns — leave whatever width/style is already set
        # (e.g. the dark margin column past the data).
        if max_length == 0:
            continue

        adjusted_width = max_length + 2  # some padding

        if adjusted_width < min_width:
            adjusted_width = min_width
        if adjusted_width > max_width:
            adjusted_width = max_width

        ws.column_dimensions[column_letter].width = adjusted_width
